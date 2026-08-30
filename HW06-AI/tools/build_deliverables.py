#!/usr/bin/env python3
"""Build the HW06 Excel workbook and the Markdown-to-PDF deliverables.

- Excel: reports/23127272_HW06_test_cases.xlsx  (Test Cases + Summary sheets)
- PDF:   output/pdf/23127272_HW06_Main_Report.pdf
         output/pdf/23127272_HW06_AI_Audit.pdf

PDF rendering uses headless Chrome/Edge (no external service).
"""
from __future__ import annotations

import csv
import json
import subprocess
import sys
from pathlib import Path

import markdown as md_lib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
OUT_PDF = ROOT / "output" / "pdf"

CHROME_CANDIDATES = [
    r"C:/Program Files/Google/Chrome/Application/chrome.exe",
    r"C:/Program Files (x86)/Google/Chrome/Application/chrome.exe",
    r"C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe",
]

CSS = """
@page { size: A4; margin: 14mm 12mm; }
* { box-sizing: border-box; }
body { font-family: 'Segoe UI', Arial, sans-serif; font-size: 11px; color: #1a1a1a; line-height: 1.45; }
h1 { font-size: 20px; border-bottom: 2px solid #333; padding-bottom: 4px; }
h2 { font-size: 15px; margin-top: 18px; border-bottom: 1px solid #bbb; padding-bottom: 2px; }
h3 { font-size: 13px; margin-top: 12px; }
code { background: #f2f2f2; padding: 1px 3px; border-radius: 3px; font-size: 10px; }
pre { background: #f6f6f6; padding: 8px; border-radius: 4px; overflow-x: auto; font-size: 9px; white-space: pre-wrap; word-wrap: break-word; }
table { width: 100%; border-collapse: collapse; margin: 8px 0; font-size: 9px; table-layout: fixed; }
th, td { border: 1px solid #ccc; padding: 4px 6px; text-align: left; vertical-align: top; word-wrap: break-word; overflow-wrap: anywhere; }
th { background: #eef2f7; }
tr { page-break-inside: avoid; }
"""


def build_excel() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    csv_path = REPORTS / "test-case-results.csv"
    with csv_path.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=(r > 1))
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
        widths = {1: 6, 2: 8, 3: 26, 4: 12, 5: 8, 6: 40, 7: 20, 8: 18, 9: 14, 10: 12, 11: 10, 12: 34, 13: 30, 14: 8, 15: 8, 16: 40, 17: 40, 18: 6}
        for idx, width in widths.items():
            if idx <= len(rows[0]):
                ws.column_dimensions[get_column_letter(idx)].width = width

    # Summary sheet
    summary = json.loads((REPORTS / "test-summary.json").read_text(encoding="utf-8"))
    ss = wb.create_sheet("Summary")
    ss.append(["Metric", "Pool A", "Pool B", "Pool C", "Total"])
    for cell in ss[1]:
        cell.fill = header_fill
        cell.font = header_font
    labels = [("total", "Designed"), ("ai_generated", "AI-generated"), ("student_added", "Student-added"),
              ("executed", "Executed"), ("passed", "Passed"), ("failed", "Failed"), ("not_run", "Not run")]
    for key, label in labels:
        ss.append([label, summary["A"][key], summary["B"][key], summary["C"][key], summary["total"][key]])
    ss.append([])
    stats = summary["newman"]
    ss.append(["Newman requests", stats["requests"]["total"]])
    ss.append(["Newman assertions", stats["assertions"]["total"]])
    ss.append(["Failed assertions", stats["assertions"]["failed"]])
    ss.append(["Confirmed bug groups", summary.get("confirmed_bug_groups", "")])
    for col in "ABCDE":
        ss.column_dimensions[col].width = 18

    out = REPORTS / "23127272_HW06_test_cases.xlsx"
    wb.save(out)
    return out


def find_browser() -> str:
    for path in CHROME_CANDIDATES:
        if Path(path).exists():
            return path
    raise RuntimeError("No Chrome/Edge found for PDF rendering")


def md_to_pdf(md_paths: list[Path], out_pdf: Path, browser: str) -> None:
    html_body = ""
    for i, md_path in enumerate(md_paths):
        text = md_path.read_text(encoding="utf-8")
        html_body += md_lib.markdown(text, extensions=["tables", "fenced_code", "sane_lists"])
        if i < len(md_paths) - 1:
            html_body += '<div style="page-break-after: always;"></div>'
    html = f"<!doctype html><html><head><meta charset='utf-8'><style>{CSS}</style></head><body>{html_body}</body></html>"
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    tmp_html = out_pdf.with_suffix(".html")
    tmp_html.write_text(html, encoding="utf-8")
    uri = "file:///" + str(tmp_html).replace("\\", "/")
    subprocess.run([
        browser, "--headless=new", "--disable-gpu", "--no-pdf-header-footer",
        "--run-all-compositor-stages-before-draw", "--virtual-time-budget=10000",
        f"--print-to-pdf={out_pdf}", uri,
    ], check=True, capture_output=True, timeout=120)
    tmp_html.unlink(missing_ok=True)


def main() -> int:
    xlsx = build_excel()
    print("Excel:", xlsx, xlsx.stat().st_size, "bytes")
    browser = find_browser()
    md_to_pdf([ROOT / "main-report.md"], OUT_PDF / "23127272_HW06_Main_Report.pdf", browser)
    md_to_pdf([ROOT / "AI docs" / "AI-Audit-Report.md", ROOT / "AI docs" / "AI_critique.md"],
              OUT_PDF / "23127272_HW06_AI_Audit.pdf", browser)
    for name in ["23127272_HW06_Main_Report.pdf", "23127272_HW06_AI_Audit.pdf"]:
        p = OUT_PDF / name
        print("PDF:", p, p.stat().st_size, "bytes" if p.exists() else "MISSING")
    return 0


if __name__ == "__main__":
    sys.exit(main())
